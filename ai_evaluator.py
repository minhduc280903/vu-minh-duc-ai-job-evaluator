"""
Phase 5: AI Job Evaluator Engine
=================================
Two-round evaluation system:
  Round 1: Fast keyword scoring (Python, instant) → all jobs
  Round 2: LLM deep analysis (Ollama) → top candidates only

Usage:
  python ai_evaluator.py                  # Run both rounds
  python ai_evaluator.py --keyword-only   # Round 1 only (instant)
  python ai_evaluator.py --llm-only       # Round 2 only (requires Round 1 done)
  python ai_evaluator.py --export         # Export results to Excel
  python ai_evaluator.py --top 30         # Show top 30 matches
  python ai_evaluator.py --reset          # Reset all scores to re-evaluate
"""

import sqlite3
import json
import asyncio
import aiohttp
import sys
import os
import re
import time
import argparse
from datetime import datetime

# === CONFIG ===
DB_FILE = "jobs.db"
PROFILE_FILE = "user_profile.json"
OLLAMA_URL = "http://localhost:11434/api/generate"
MODEL_NAME = "qwen2.5:14b"
LLM_TIMEOUT = 120  # seconds per job
LLM_THRESHOLD = 25  # minimum keyword score to send to LLM
BATCH_SIZE = 1  # sequential for stability with 14b model
EXPORT_FILE = "job_matches_{date}.xlsx"

# === ENCODING FIX (Windows) ===
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")


def load_profile():
    with open(PROFILE_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_columns(conn):
    """Add evaluation columns if they don't exist."""
    c = conn.cursor()
    # Check existing columns
    c.execute("PRAGMA table_info(jobs)")
    existing = {row[1] for row in c.fetchall()}

    new_cols = {
        "keyword_score": "INTEGER DEFAULT -1",
        "llm_score": "INTEGER DEFAULT -1",
        "llm_rationale": "TEXT",
        "llm_pros": "TEXT",
        "llm_cons": "TEXT",
        "final_score": "INTEGER DEFAULT -1",
    }
    for col, typedef in new_cols.items():
        if col not in existing:
            c.execute(f"ALTER TABLE jobs ADD COLUMN {col} {typedef}")
            print(f"  Added column: {col}")
    conn.commit()


# ============================================================
# ROUND 1: KEYWORD SCORING (instant, no LLM)
# ============================================================

def keyword_score_job(job, profile):
    """Score a single job based on keyword matching. Returns 0-100."""
    title = (job["title"] or "").lower()
    company = (job["company"] or "").lower()
    desc = (job["description"] or "").lower()
    reqs = (job["requirements"] or "").lower()
    skills_field = (job["skills"] or "").lower()
    benefits = (job["benefits"] or "").lower()
    location = (job["location"] or "").lower()
    level = (job["level"] or "").lower()

    # Combine text fields for searching
    all_text = f"{desc} {reqs} {skills_field} {benefits}"
    title_company = f"{title} {company}"

    score = 0

    # --- 0. IRRELEVANT TITLE PENALTY (early exit) ---
    # These roles are definitely NOT what the candidate wants
    irrelevant_titles = [
        # Frontend / Mobile
        "frontend", "front-end", "react native", "ios dev", "android dev",
        "mobile dev", "flutter", "swift dev", "kotlin dev", "react dev",
        "angular", "vue.js dev", "nextjs",
        # Backend / Infra (not data-related)
        "backend", "back-end", "fullstack", "full-stack", "full stack",
        "devops", "sre ", "site reliability", "infrastructure",
        "platform engineer", "cloud engineer", "system admin",
        "network engineer", "security engineer", "penetration tester",
        # UI/UX / Design / Content / Marketing
        "ui/ux", "ux designer", "ui designer", "graphic design", "content",
        "marketing", "seo ", "social media", "copywriter",
        # Sales / Customer-facing
        "sales", "telesales", "tư vấn bán", "account manager",
        "customer success", "chăm sóc khách", "tư vấn viên",
        # Game / Embedded / Hardware
        "game dev", "game design", "unity dev", "unreal",
        "embedded", "firmware", "hardware",
        # Education / HR / Admin
        "teacher", "giáo viên", "gia sư", "giảng dạy",
        "hr ", "nhân sự", "hành chính", "admin", "receptionist",
        "lễ tân", "thư ký",
        # QA / Testing (not QA analyst in finance)
        "manual tester", "manual qa", "qa engineer", "qc engineer",
        "qa automation", "quality assurance", "test engineer",
        "qa manager", "tester",
        # PM / Scrum / Delivery
        "project manager", "scrum master", "delivery manager",
        "product owner", "technical lead", "tech lead",
        # Java / .NET / PHP / Other pure dev
        "java developer", ".net developer", "php developer",
        "c# developer", "c++ developer", "ruby developer",
        "golang developer", "go developer", "rust developer",
        "nodejs developer", "node.js developer",
        "software engineer", "lập trình viên",
    ]
    for kw in irrelevant_titles:
        if kw in title:
            return 0  # instant reject

    # --- 1. Title Match (max 25) ---
    title_score = 0
    for tier_name, tier in profile["title_keywords"].items():
        if tier_name.startswith("_"):  # skip _comment
            continue
        if not isinstance(tier, dict) or "keywords" not in tier:
            continue
        for kw in tier["keywords"]:
            if kw.lower() in title:
                title_score = max(title_score, tier["score"])
        if title_score > 0:
            break
    score += title_score

    # --- 2. Skill Match (max 25) ---
    skill_score = 0
    for level in ["high_value", "medium_value", "low_value"]:
        group = profile["skill_keywords"][level]
        for kw in group["keywords"]:
            if kw.lower() in all_text or kw.lower() in skills_field:
                skill_score += group["points_each"]
    score += min(skill_score, 25)

    # --- 3. Industry Match (max 25) ---
    industry_score = 0
    for tier_name in ["tier_s_finance", "tier_a_finance_domain", "tier_b_tech"]:
        tier = profile["industry_keywords"][tier_name]
        for kw in tier["keywords"]:
            if kw.lower() in title_company or kw.lower() in all_text:
                industry_score = max(industry_score, tier["points"])
                break
        if industry_score >= 20:
            break
    score += industry_score

    # --- 4. Work Style (max 15, can go negative) ---
    style_score = 0
    for kw in profile["work_style"]["positive"]["keywords"]:
        if kw.lower() in all_text:
            style_score += profile["work_style"]["positive"]["points_each"]
    for kw in profile["work_style"]["negative"]["keywords"]:
        if kw.lower() in all_text:
            style_score += profile["work_style"]["negative"]["points_each"]
    score += max(min(style_score, 15), -10)

    # --- 5. Experience Level (max 10, can go very negative) ---
    exp_score = 0
    exp_text = f"{title} {reqs} {level}"
    
    # 5a. STRUCTURED experience from level field (strongest signal)
    if level:
        import re as _re
        # Extract years from level field like "3 - 5 Năm", "5 Năm", "3-5 years"
        year_match = _re.search(r'(\d+)\s*[-–]\s*(\d+)', level)
        if year_match:
            min_years = int(year_match.group(1))
            max_years = int(year_match.group(2))
        else:
            single_match = _re.search(r'(\d+)', level)
            if single_match:
                min_years = int(single_match.group(1))
                max_years = min_years
            else:
                min_years = 0
                max_years = 0
        
        if min_years >= 5:
            exp_score = -25  # Way too senior
        elif min_years >= 3:
            exp_score = -15  # Senior, very hard for fresher
        elif min_years >= 2:
            exp_score = -5   # Stretch but possible
        elif min_years <= 1:
            exp_score = 10   # Ideal for fresher
    
    # 5b. Title-based seniority check
    if 'giám đốc' in title or 'director' in title or 'head of' in title or 'trưởng phòng' in title:
        exp_score = min(exp_score, -20)  # Leadership roles = not for fresher
    elif 'senior' in title or 'lead' in title or 'principal' in title or 'staff' in title:
        exp_score = min(exp_score, -10)
    
    # 5c. Fallback: text-based experience keywords (if no structured level)
    if not level and exp_score == 0:
        for level_name in ["ideal", "acceptable", "stretch"]:
            level_cfg = profile["experience_level"][level_name]
            for kw in level_cfg["keywords"]:
                if kw.lower() in exp_text:
                    exp_score = level_cfg["points"]
                    break
            if exp_score != 0:
                break
    
    score += exp_score

    # --- 6. Location Bonus (Hanoi +5) ---
    hanoi_keywords = ["hà nội", "ha noi", "hanoi", "cầu giấy", "nam từ liêm",
                      "thanh xuân", "đống đa", "hoàn kiếm", "ba đình"]
    for kw in hanoi_keywords:
        if kw in location or kw in title:
            score += 5
            break

    # --- 7. Title has NO relevant keyword at all → cap at 20 ---
    # If the title doesn't match any tier, this is likely not relevant
    if title_score == 0:
        score = min(score, 15)

    return max(0, min(100, score))


def run_keyword_scoring(conn, profile):
    """Score ALL jobs with keyword matching."""
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM jobs")
    total = c.fetchone()[0]

    c.execute("""SELECT id, title, company, description, requirements,
                        skills, benefits, location, level
                 FROM jobs""")
    jobs = c.fetchall()

    print(f"\n{'='*60}")
    print(f"  ROUND 1: KEYWORD SCORING — {total} jobs")
    print(f"{'='*60}")

    scored = 0
    distribution = {"0-9": 0, "10-24": 0, "25-49": 0, "50-74": 0, "75-100": 0}

    for job in jobs:
        ks = keyword_score_job(job, profile)
        c.execute(
            "UPDATE jobs SET keyword_score = ? WHERE id = ?",
            (ks, job["id"]),
        )
        scored += 1

        if ks < 10:
            distribution["0-9"] += 1
        elif ks < 25:
            distribution["10-24"] += 1
        elif ks < 50:
            distribution["25-49"] += 1
        elif ks < 75:
            distribution["50-74"] += 1
        else:
            distribution["75-100"] += 1

        if scored % 500 == 0:
            print(f"  Scored {scored}/{total}...")
            conn.commit()

    conn.commit()

    print(f"\n  Done! Scored {scored} jobs.")
    print(f"\n  Distribution:")
    for bucket, count in distribution.items():
        bar = "#" * (count // 10)
        print(f"    {bucket:>6}: {count:>5}  {bar}")

    # Show candidates for LLM
    c.execute(
        "SELECT COUNT(*) FROM jobs WHERE keyword_score >= ?", (LLM_THRESHOLD,)
    )
    llm_candidates = c.fetchone()[0]
    print(f"\n  LLM candidates (score >= {LLM_THRESHOLD}): {llm_candidates}")

    return scored


# ============================================================
# ROUND 2: LLM DEEP EVALUATION (Ollama)
# ============================================================

LLM_PROMPT_TEMPLATE = """{system_prompt}

JOB POSTING:
Platform: {platform}
Title: {title}
Company: {company}
Skills: {skills}
Location: {location}
Salary: {salary}
Experience Required: {level}

Description:
{description}

Requirements:
{requirements}

Benefits:
{benefits}

---
Evaluate this job for the candidate. Output ONLY valid JSON (no markdown, no codeblocks):
{{"score": <0-100>, "rationale": "<2-3 sentences in Vietnamese>", "pros": ["<point1>", "<point2>"], "cons": ["<point1>", "<point2>"]}}"""


async def llm_evaluate_job(session, job, system_prompt):
    """Send a single job to Ollama for deep evaluation."""
    prompt = LLM_PROMPT_TEMPLATE.format(
        system_prompt=system_prompt,
        platform=job["platform"] or "",
        title=job["title"] or "",
        company=job["company"] or "",
        skills=job["skills"] or "",
        location=job["location"] or "",
        salary=job["salary"] or "N/A",
        level=job["level"] or "Không rõ",
        description=(job["description"] or "")[:2000],
        requirements=(job["requirements"] or "")[:1500],
        benefits=(job["benefits"] or "")[:800],
    )

    payload = {
        "model": MODEL_NAME,
        "prompt": prompt,
        "stream": False,
        "format": "json",
        "options": {"temperature": 0.1, "num_predict": 512},
    }

    try:
        timeout = aiohttp.ClientTimeout(total=LLM_TIMEOUT)
        async with session.post(OLLAMA_URL, json=payload, timeout=timeout) as resp:
            data = await resp.json()
            result = json.loads(data["response"])
            # Validate
            score = int(result.get("score", 0))
            score = max(0, min(100, score))
            return {
                "score": score,
                "rationale": result.get("rationale", ""),
                "pros": json.dumps(result.get("pros", []), ensure_ascii=False),
                "cons": json.dumps(result.get("cons", []), ensure_ascii=False),
            }
    except asyncio.TimeoutError:
        return {"score": -1, "rationale": "Timeout", "pros": "[]", "cons": "[]"}
    except Exception as e:
        return {
            "score": -1,
            "rationale": f"Error: {str(e)[:200]}",
            "pros": "[]",
            "cons": "[]",
        }


async def run_llm_evaluation(conn, profile):
    """Evaluate top keyword-scored jobs with Ollama LLM."""
    c = conn.cursor()

    # Get jobs that passed keyword filter but haven't been LLM-scored
    c.execute(
        """SELECT id, platform, title, company, description, requirements,
                  benefits, skills, location, salary, keyword_score, level
           FROM jobs
           WHERE keyword_score >= ? AND llm_score = -1
           ORDER BY keyword_score DESC""",
        (LLM_THRESHOLD,),
    )
    jobs = c.fetchall()

    if not jobs:
        print("\n  No jobs pending LLM evaluation.")
        c.execute(
            "SELECT COUNT(*) FROM jobs WHERE keyword_score >= ? AND llm_score > -1",
            (LLM_THRESHOLD,),
        )
        done = c.fetchone()[0]
        print(f"  Already evaluated: {done}")
        return 0

    total = len(jobs)
    system_prompt = profile.get("llm_evaluation_prompt", "")

    print(f"\n{'='*60}")
    print(f"  ROUND 2: LLM DEEP EVALUATION — {total} jobs")
    print(f"  Model: {MODEL_NAME}")
    print(f"  Batch size: {BATCH_SIZE}")
    print(f"{'='*60}")

    # Check Ollama is running
    try:
        async with aiohttp.ClientSession() as test_session:
            async with test_session.get(
                "http://localhost:11434/api/tags", timeout=aiohttp.ClientTimeout(total=5)
            ) as resp:
                if resp.status != 200:
                    print("\n  ERROR: Ollama not responding. Start with: ollama serve")
                    return 0
    except Exception:
        print("\n  ERROR: Cannot connect to Ollama at localhost:11434")
        print("  Start Ollama with: ollama serve")
        return 0

    evaluated = 0
    errors = 0
    start_time = time.time()

    async with aiohttp.ClientSession() as session:
        # Process in batches
        for i in range(0, total, BATCH_SIZE):
            batch = jobs[i : i + BATCH_SIZE]
            tasks = [
                llm_evaluate_job(session, job, system_prompt) for job in batch
            ]
            results = await asyncio.gather(*tasks)

            for job, result in zip(batch, results):
                if result["score"] >= 0:
                    # Calculate final score: 40% keyword + 60% LLM
                    final = int(job["keyword_score"] * 0.4 + result["score"] * 0.6)
                    c.execute(
                        """UPDATE jobs SET
                           llm_score = ?, llm_rationale = ?,
                           llm_pros = ?, llm_cons = ?,
                           final_score = ?,
                           relevance_score = ?
                           WHERE id = ?""",
                        (
                            result["score"],
                            result["rationale"],
                            result["pros"],
                            result["cons"],
                            final,
                            final,  # also update the original relevance_score
                            job["id"],
                        ),
                    )
                    evaluated += 1
                else:
                    errors += 1

            conn.commit()

            # Progress
            elapsed = time.time() - start_time
            rate = evaluated / elapsed if elapsed > 0 else 0
            eta = (total - i - len(batch)) / rate if rate > 0 else 0
            print(
                f"  [{evaluated + errors}/{total}] "
                f"Evaluated: {evaluated} | Errors: {errors} | "
                f"Speed: {rate:.1f} jobs/s | ETA: {eta/60:.0f}m"
            )

    elapsed = time.time() - start_time
    print(f"\n  Done! {evaluated} evaluated, {errors} errors in {elapsed/60:.1f}m")
    return evaluated


# ============================================================
# RESULTS & EXPORT
# ============================================================

def show_top_matches(conn, n=30):
    """Display top N matched jobs."""
    c = conn.cursor()

    # Determine which score to use
    c.execute("SELECT COUNT(*) FROM jobs WHERE final_score > -1")
    has_final = c.fetchone()[0] > 0

    if has_final:
        score_col = "final_score"
        label = "Final Score (40% keyword + 60% LLM)"
    else:
        score_col = "keyword_score"
        label = "Keyword Score (LLM not yet run)"

    c.execute(
        f"""SELECT platform, title, company, url, salary, location,
                   keyword_score, llm_score, {score_col} as score,
                   llm_rationale, llm_pros, llm_cons
            FROM jobs
            WHERE {score_col} > 0
            ORDER BY {score_col} DESC
            LIMIT ?""",
        (n,),
    )
    jobs = c.fetchall()

    if not jobs:
        print("\n  No scored jobs found. Run keyword scoring first.")
        return

    print(f"\n{'='*80}")
    print(f"  TOP {n} JOB MATCHES — {label}")
    print(f"{'='*80}")

    for i, job in enumerate(jobs, 1):
        score = job["score"]

        # Color indicator
        if score >= 75:
            tier = "S"
        elif score >= 50:
            tier = "A"
        elif score >= 35:
            tier = "B"
        else:
            tier = "C"

        print(f"\n  #{i:>2} [{tier}] Score: {score}/100 (KW:{job['keyword_score']} | LLM:{job['llm_score']})")
        print(f"      {job['title']}")
        print(f"      {job['company']} | {job['platform']} | {job['location'] or 'N/A'}")
        if job["salary"]:
            print(f"      Salary: {job['salary']}")
        if job["url"]:
            print(f"      URL: {job['url']}")
        if job["llm_rationale"]:
            print(f"      AI: {job['llm_rationale']}")
        if job["llm_pros"] and job["llm_pros"] != "[]":
            try:
                pros = json.loads(job["llm_pros"])
                if pros:
                    print(f"      +  {' | '.join(pros)}")
            except json.JSONDecodeError:
                pass
        if job["llm_cons"] and job["llm_cons"] != "[]":
            try:
                cons = json.loads(job["llm_cons"])
                if cons:
                    print(f"      -  {' | '.join(cons)}")
            except json.JSONDecodeError:
                pass


def export_to_excel(conn):
    """Export all scored jobs to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("\n  openpyxl not installed. Installing...")
        os.system(f"{sys.executable} -m pip install openpyxl")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    c = conn.cursor()

    # Determine score column
    c.execute("SELECT COUNT(*) FROM jobs WHERE final_score > -1")
    has_final = c.fetchone()[0] > 0
    score_col = "final_score" if has_final else "keyword_score"

    c.execute(
        f"""SELECT platform, title, company, url, salary, location,
                   keyword_score, llm_score, {score_col} as score,
                   llm_rationale, llm_pros, llm_cons, skills, deadline,
                   description, requirements
            FROM jobs
            WHERE {score_col} > 0
            ORDER BY {score_col} DESC"""
    )
    jobs = c.fetchall()

    if not jobs:
        print("  No scored jobs to export.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Matches"

    # === HEADER STYLING ===
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # Tier colors
    tier_colors = {
        "S": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        "A": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "B": PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
        "C": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    }

    headers = [
        "#", "Tier", "Score", "KW", "LLM", "Platform",
        "Title", "Company", "Salary", "Location",
        "Description", "Requirements", "Skills", "AI Rationale", "Pros", "Cons", "URL", "Deadline",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # === DATA ROWS ===
    for i, job in enumerate(jobs, 1):
        score = job["score"]
        tier = "S" if score >= 75 else "A" if score >= 50 else "B" if score >= 35 else "C"

        row_data = [
            i,
            tier,
            score,
            job["keyword_score"],
            job["llm_score"] if job["llm_score"] > -1 else "—",
            job["platform"],
            job["title"],
            job["company"],
            job["salary"] or "—",
            job["location"] or "—",
            (job["description"] or "—")[:500],
            (job["requirements"] or "—")[:500],
            job["skills"] or "—",
            job["llm_rationale"] or "—",
            _format_json_list(job["llm_pros"]),
            _format_json_list(job["llm_cons"]),
            job["url"] or "",
            job["deadline"] or "—",
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.border = thin_border
            cell.fill = tier_colors.get(tier, PatternFill())
            cell.alignment = Alignment(
                vertical="center", wrap_text=(col in [7, 11, 12, 13, 14, 15, 16])
            )

    # === COLUMN WIDTHS ===
    widths = {
        1: 4, 2: 5, 3: 6, 4: 5, 5: 5, 6: 12,
        7: 50, 8: 25, 9: 18, 10: 15,
        11: 40, 12: 40, 13: 30, 14: 45, 15: 30, 16: 30, 17: 40, 18: 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Auto-filter
    ws.auto_filter.ref = f"A1:R{len(jobs) + 1}"
    ws.freeze_panes = "A2"

    # === SUMMARY SHEET ===
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Job Match Summary"
    ws2["A1"].font = Font(size=14, bold=True)
    ws2["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws2["A4"] = f"Total jobs in DB: {_count_total(conn)}"
    ws2["A5"] = f"Jobs with scores > 0: {len(jobs)}"
    ws2["A7"] = "Tier Distribution:"
    ws2["A8"] = f"  Tier S (75-100): {sum(1 for j in jobs if j['score'] >= 75)}"
    ws2["A9"] = f"  Tier A (50-74):  {sum(1 for j in jobs if 50 <= j['score'] < 75)}"
    ws2["A10"] = f"  Tier B (35-49):  {sum(1 for j in jobs if 35 <= j['score'] < 50)}"
    ws2["A11"] = f"  Tier C (1-34):   {sum(1 for j in jobs if 0 < j['score'] < 35)}"

    filename = EXPORT_FILE.format(date=datetime.now().strftime("%Y%m%d_%H%M"))
    wb.save(filename)
    print(f"\n  Exported {len(jobs)} jobs to: {filename}")
    print(f"  Tier S: {sum(1 for j in jobs if j['score'] >= 75)}")
    print(f"  Tier A: {sum(1 for j in jobs if 50 <= j['score'] < 75)}")
    print(f"  Tier B: {sum(1 for j in jobs if 35 <= j['score'] < 50)}")
    return filename


def _format_json_list(s):
    if not s or s == "[]":
        return "—"
    try:
        items = json.loads(s)
        return " | ".join(items) if items else "—"
    except (json.JSONDecodeError, TypeError):
        return str(s)


def _count_total(conn):
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM jobs")
    return c.fetchone()[0]


def reset_scores(conn):
    """Reset all evaluation scores."""
    c = conn.cursor()
    c.execute(
        """UPDATE jobs SET
           keyword_score = -1, llm_score = -1, final_score = -1,
           llm_rationale = NULL, llm_pros = NULL, llm_cons = NULL,
           relevance_score = -1, evaluation_reason = NULL"""
    )
    conn.commit()
    print(f"  Reset all scores for {c.rowcount} jobs.")


def show_stats(conn):
    """Show current evaluation statistics."""
    c = conn.cursor()

    c.execute("SELECT COUNT(*) FROM jobs")
    total = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM jobs WHERE keyword_score > -1")
    kw_done = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM jobs WHERE llm_score > -1")
    llm_done = c.fetchone()[0]

    c.execute(
        "SELECT COUNT(*) FROM jobs WHERE keyword_score >= ?", (LLM_THRESHOLD,)
    )
    llm_candidates = c.fetchone()[0]

    print(f"\n  === EVALUATION STATUS ===")
    print(f"  Total jobs:         {total}")
    print(f"  Keyword scored:     {kw_done}/{total}")
    print(f"  LLM candidates:     {llm_candidates} (score >= {LLM_THRESHOLD})")
    print(f"  LLM evaluated:      {llm_done}/{llm_candidates}")
    print(f"  Pending LLM:        {llm_candidates - llm_done}")

    # Top 5 preview
    c.execute(
        """SELECT title, company, keyword_score FROM jobs
           WHERE keyword_score > -1 ORDER BY keyword_score DESC LIMIT 5"""
    )
    preview = c.fetchall()
    if preview:
        print(f"\n  Top 5 by keyword score:")
        for j in preview:
            print(f"    [{j['keyword_score']:>3}] {j['company']} — {j['title'][:60]}")


# ============================================================
# MAIN
# ============================================================

async def main():
    parser = argparse.ArgumentParser(description="AI Job Evaluator Engine")
    parser.add_argument("--keyword-only", action="store_true", help="Run keyword scoring only")
    parser.add_argument("--llm-only", action="store_true", help="Run LLM evaluation only")
    parser.add_argument("--export", action="store_true", help="Export results to Excel")
    parser.add_argument("--top", type=int, default=0, help="Show top N matches")
    parser.add_argument("--reset", action="store_true", help="Reset all scores")
    parser.add_argument("--stats", action="store_true", help="Show evaluation stats")
    parser.add_argument("--model", type=str, default=MODEL_NAME, help="Ollama model name")
    parser.add_argument("--threshold", type=int, default=LLM_THRESHOLD, help="Min keyword score for LLM")
    parser.add_argument("--batch", type=int, default=BATCH_SIZE, help="LLM batch size")
    args = parser.parse_args()

    # Update config from args
    _update_config(args.model, args.threshold, args.batch)

    print("\n  ╔══════════════════════════════════════════╗")
    print("  ║   AI JOB EVALUATOR ENGINE — Phase 5     ║")
    print("  ╚══════════════════════════════════════════╝")

    conn = get_db()
    ensure_columns(conn)
    profile = load_profile()

    if args.reset:
        reset_scores(conn)
        conn.close()
        return

    if args.stats:
        show_stats(conn)
        conn.close()
        return

    if args.export:
        export_to_excel(conn)
        conn.close()
        return

    if args.top:
        show_top_matches(conn, args.top)
        conn.close()
        return

    # Default: run full pipeline
    if not args.llm_only:
        run_keyword_scoring(conn, profile)
        show_stats(conn)

    if not args.keyword_only:
        await run_llm_evaluation(conn, profile)

    # Always show top results
    show_top_matches(conn, 30)

    # Always export
    export_to_excel(conn)

    conn.close()
    print("\n  All done!")


def _update_config(model, threshold, batch):
    global MODEL_NAME, LLM_THRESHOLD, BATCH_SIZE
    MODEL_NAME = model
    LLM_THRESHOLD = threshold
    BATCH_SIZE = batch


if __name__ == "__main__":
    asyncio.run(main())
